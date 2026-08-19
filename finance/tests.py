from datetime import date
from decimal import Decimal
from io import BytesIO
import re
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.conf import settings
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import IntegrityError, transaction
from django.test import TestCase, TransactionTestCase
from django.urls import reverse
from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError

from customers.models import Customer
from inventory.models import (
    Category,
    Product,
    StockBalance,
    StockBatch,
    StockMovement,
    Unit,
    Warehouse,
    WasteLoss,
    WasteLossItem,
)
from inventory.services import confirm_waste_loss
from pos.models import POSPayment, POSSale, POSSaleItem
from pos.services import complete_sale
from purchasing.models import PurchaseInvoice, PurchaseInvoiceItem
from purchasing.services import confirm_purchase
from suppliers.models import Supplier

from .forms import AccountForm
from .models import (
    Account,
    FiscalPeriod,
    FiscalPeriodAction,
    FiscalYear,
    FinanceAuditLog,
    JournalEntry,
    JournalEntryLine,
    OpenItem,
    OpenItemAllocation,
    PaymentVoucher,
    PeriodStatus,
    ReceiptVoucher,
)
from .reports import (
    generate_balance_sheet,
    generate_cash_flow_statement,
    generate_customer_statement,
    generate_income_statement,
    generate_payables_aging,
    generate_receivables_aging,
    generate_supplier_statement,
    get_account_balance,
)
from .services import (
    create_fiscal_year,
    get_period_summary,
    get_posting_account,
    post_journal_entry,
    post_payment_voucher,
    post_pos_sale,
    post_purchase_invoice,
    post_receipt_voucher,
    reverse_journal_entry,
    set_fiscal_year_status,
    set_period_status,
)


def create_required_fiscal_years(*years):
    """Create each calendar year once for tests that post journal entries."""
    for year in set(years):
        create_fiscal_year(year)


class AccountHierarchyViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="chart-accountant",
            password="test-password",
            role="accountant",
        )
        self.client.force_login(self.user)
        self.assets = Account.objects.create(
            code="1000",
            name="Assets",
            account_type=Account.AccountType.ASSET,
            allow_posting=False,
        )
        self.current_assets = Account.objects.create(
            code="1100",
            name="Current assets",
            account_type=Account.AccountType.ASSET,
            parent=self.assets,
            allow_posting=False,
        )
        self.cash = Account.objects.create(
            code="1110",
            name="Cash",
            account_type=Account.AccountType.ASSET,
            parent=self.current_assets,
            is_cash_equivalent=True,
        )
        self.revenue = Account.objects.create(
            code="4000",
            name="Revenue",
            account_type=Account.AccountType.REVENUE,
            allow_posting=False,
        )

    def test_account_list_builds_parent_child_tree(self):
        response = self.client.get(reverse("finance:account_list"))

        self.assertEqual(response.status_code, 200)
        rows = response.context["account_rows"]
        self.assertEqual(
            [(row["account"].code, row["depth"]) for row in rows],
            [("1000", 0), ("1100", 1), ("1110", 2), ("4000", 0)],
        )
        self.assertTrue(rows[0]["has_children"])
        self.assertFalse(rows[2]["has_children"])
        self.assertContains(response, reverse("finance:account_create"))
        self.assertContains(
            response,
            reverse("finance:account_update", args=[self.cash.pk]),
        )

    def test_accountant_can_create_posting_child_account(self):
        response = self.client.post(
            reverse("finance:account_create"),
            {
                "code": "1120",
                "name": "Bank",
                "account_type": Account.AccountType.ASSET,
                "parent": self.current_assets.pk,
                "allow_posting": "on",
                "is_cash_equivalent": "on",
                "is_active": "on",
            },
        )

        account = Account.objects.get(code="1120")
        self.assertRedirects(response, reverse("finance:account_list"))
        self.assertEqual(account.parent, self.current_assets)
        self.assertTrue(account.allow_posting)
        self.assertTrue(account.is_cash_equivalent)

    def test_accountant_can_edit_account(self):
        response = self.client.post(
            reverse("finance:account_update", args=[self.cash.pk]),
            {
                "code": self.cash.code,
                "name": "Main cash",
                "account_type": Account.AccountType.ASSET,
                "parent": self.current_assets.pk,
                "allow_posting": "on",
                "is_cash_equivalent": "on",
                "is_active": "on",
            },
        )

        self.assertRedirects(response, reverse("finance:account_list"))
        self.cash.refresh_from_db()
        self.assertEqual(self.cash.name, "Main cash")

    def test_form_rejects_posting_parent_and_different_account_type(self):
        posting_parent_form = AccountForm(
            data={
                "code": "1125",
                "name": "Petty cash",
                "account_type": Account.AccountType.ASSET,
                "parent": self.cash.pk,
                "allow_posting": "on",
                "is_active": "on",
            }
        )
        wrong_type_form = AccountForm(
            data={
                "code": "4100",
                "name": "Wrong child",
                "account_type": Account.AccountType.ASSET,
                "parent": self.revenue.pk,
                "allow_posting": "on",
                "is_active": "on",
            }
        )

        self.assertFalse(posting_parent_form.is_valid())
        self.assertIn("parent", posting_parent_form.errors)
        self.assertFalse(wrong_type_form.is_valid())
        self.assertIn("parent", wrong_type_form.errors)

    def test_invalid_account_response_contains_toast_detectable_field_error(self):
        response = self.client.post(
            reverse("finance:account_create"),
            {
                "code": "1125",
                "name": "Petty cash",
                "account_type": Account.AccountType.ASSET,
                "parent": self.cash.pk,
                "allow_posting": "on",
                "is_active": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="form-error"')
        self.assertContains(
            response,
            "A posting account cannot be used as a parent account.",
        )
        self.assertFalse(Account.objects.filter(code="1125").exists())

    def test_form_rejects_cycle_and_posting_summary_account(self):
        cycle_form = AccountForm(
            data={
                "code": self.assets.code,
                "name": self.assets.name,
                "account_type": Account.AccountType.ASSET,
                "parent": self.current_assets.pk,
                "is_active": "on",
            },
            instance=self.assets,
        )
        posting_summary_form = AccountForm(
            data={
                "code": self.current_assets.code,
                "name": self.current_assets.name,
                "account_type": Account.AccountType.ASSET,
                "parent": self.assets.pk,
                "allow_posting": "on",
                "is_active": "on",
            },
            instance=self.current_assets,
        )

        self.assertFalse(cycle_form.is_valid())
        self.assertIn("parent", cycle_form.errors)
        self.assertFalse(posting_summary_form.is_valid())
        self.assertIn("allow_posting", posting_summary_form.errors)

    def test_cash_equivalent_must_be_posting_asset(self):
        form = AccountForm(
            data={
                "code": "4100",
                "name": "Cash-like revenue",
                "account_type": Account.AccountType.REVENUE,
                "parent": self.revenue.pk,
                "is_cash_equivalent": "on",
                "is_active": "on",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("is_cash_equivalent", form.errors)

    def test_parent_type_cannot_change_while_children_keep_old_type(self):
        form = AccountForm(
            data={
                "code": self.assets.code,
                "name": self.assets.name,
                "account_type": Account.AccountType.LIABILITY,
                "is_active": "on",
            },
            instance=self.assets,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("account_type", form.errors)


class FinanceAuditLogTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="audit-accountant",
            password="test-password",
            role="accountant",
        )
        self.client.force_login(self.user)

    def test_account_create_and_update_record_before_and_after_values(self):
        create_response = self.client.post(
            reverse("finance:account_create"),
            {
                "code": "AUD1000",
                "name": "Audit cash",
                "account_type": Account.AccountType.ASSET,
                "allow_posting": "on",
                "is_cash_equivalent": "on",
                "is_active": "on",
            },
        )
        account = Account.objects.get(code="AUD1000")
        created_log = FinanceAuditLog.objects.get(
            entity_type="finance.account",
            object_id=str(account.pk),
            action=FinanceAuditLog.Action.CREATED,
        )

        self.assertRedirects(create_response, reverse("finance:account_list"))
        self.assertEqual(created_log.actor, self.user)
        self.assertEqual(created_log.changes["name"]["after"], "Audit cash")

        update_response = self.client.post(
            reverse("finance:account_update", args=[account.pk]),
            {
                "code": account.code,
                "name": "Main audit cash",
                "account_type": Account.AccountType.ASSET,
                "allow_posting": "on",
                "is_cash_equivalent": "on",
                "is_active": "on",
            },
        )
        updated_log = FinanceAuditLog.objects.get(
            entity_type="finance.account",
            object_id=str(account.pk),
            action=FinanceAuditLog.Action.UPDATED,
        )

        self.assertRedirects(update_response, reverse("finance:account_list"))
        self.assertEqual(
            updated_log.changes["name"],
            {"before": "Audit cash", "after": "Main audit cash"},
        )
        self.assertEqual(set(updated_log.changes), {"name"})

    def test_successful_post_is_logged_and_failed_post_is_not(self):
        create_fiscal_year(2026, user=self.user)
        cash = Account.objects.create(
            code="AUD1100",
            name="Audit posting cash",
            account_type=Account.AccountType.ASSET,
        )
        capital = Account.objects.create(
            code="AUD3100",
            name="Audit posting capital",
            account_type=Account.AccountType.EQUITY,
        )
        journal = JournalEntry.objects.create(
            entry_number="AUD-JE-001",
            date=date(2026, 8, 19),
            description="Audit posting",
            created_by=self.user,
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=cash,
            debit=Decimal("25.000"),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=capital,
            credit=Decimal("25.000"),
        )

        post_journal_entry(journal.pk, self.user)

        posted_log = FinanceAuditLog.objects.get(
            entity_type="finance.journalentry",
            object_id=str(journal.pk),
            action=FinanceAuditLog.Action.POSTED,
        )
        self.assertEqual(posted_log.actor, self.user)
        self.assertEqual(posted_log.changes["total_debit"]["after"], "25.000")

        invalid = JournalEntry.objects.create(
            entry_number="AUD-JE-002",
            date=date(2026, 8, 19),
            created_by=self.user,
        )
        with self.assertRaises(ValidationError):
            post_journal_entry(invalid.pk, self.user)
        self.assertFalse(
            FinanceAuditLog.objects.filter(
                entity_type="finance.journalentry",
                object_id=str(invalid.pk),
                action=FinanceAuditLog.Action.POSTED,
            ).exists()
        )

    def test_audit_page_is_filterable_and_paginated(self):
        FinanceAuditLog.objects.bulk_create(
            [
                FinanceAuditLog(
                    actor=self.user,
                    actor_label=self.user.username,
                    action=(
                        FinanceAuditLog.Action.CREATED
                        if index % 2
                        else FinanceAuditLog.Action.UPDATED
                    ),
                    entity_type="finance.account",
                    entity_label="Account",
                    object_id=str(index),
                    object_repr=f"AUD-{index:03d}",
                    changes={},
                )
                for index in range(25)
            ]
        )

        first_page = self.client.get(reverse("finance:audit_log"))
        second_page = self.client.get(reverse("finance:audit_log"), {"page": 2})
        third_page = self.client.get(reverse("finance:audit_log"), {"page": 3})
        filtered = self.client.get(
            reverse("finance:audit_log"),
            {"action": FinanceAuditLog.Action.CREATED, "q": "AUD-001"},
        )

        self.assertEqual(len(first_page.context["page_obj"]), 10)
        self.assertEqual(len(second_page.context["page_obj"]), 10)
        self.assertEqual(len(third_page.context["page_obj"]), 5)
        self.assertEqual(first_page.context["page_obj"].paginator.num_pages, 3)
        self.assertEqual(filtered.context["page_obj"].paginator.count, 1)
        self.assertContains(first_page, "Audit records are read-only.")

    def test_ajax_audit_request_returns_only_dynamic_results(self):
        FinanceAuditLog.objects.create(
            actor=self.user,
            actor_label=self.user.username,
            action=FinanceAuditLog.Action.UPDATED,
            entity_type="finance.account",
            entity_label="Account",
            object_id="42",
            object_repr="AJAX account",
            changes={},
        )

        response = self.client.get(
            reverse("finance:audit_log"),
            {"q": "AJAX account"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "finance/partials/audit_log_results.html")
        self.assertContains(response, "AJAX account")
        self.assertNotContains(response, "audit-filter-form")

    def test_audit_log_exports_filtered_excel_with_change_details(self):
        FinanceAuditLog.objects.create(
            actor=self.user,
            actor_label=self.user.username,
            action=FinanceAuditLog.Action.CREATED,
            entity_type="finance.account",
            entity_label="Account",
            object_id="51",
            object_repr="Audit export account",
            changes={"name": {"before": "Old name", "after": "New name"}},
        )
        FinanceAuditLog.objects.create(
            actor=self.user,
            actor_label=self.user.username,
            action=FinanceAuditLog.Action.UPDATED,
            entity_type="finance.account",
            entity_label="Account",
            object_id="52",
            object_repr="Excluded account",
            changes={},
        )

        response = self.client.get(
            reverse("finance:audit_log"),
            {
                "q": "Audit export",
                "action": FinanceAuditLog.Action.CREATED,
                "export": "xlsx",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("finance-audit-log.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        values = "\n".join(
            str(value)
            for row in workbook.active.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        self.assertIn("Audit export account", values)
        self.assertIn("Old name → New name", values)
        self.assertNotIn("Excluded account", values)

    def test_audit_page_shows_excel_and_pdf_exports(self):
        response = self.client.get(
            reverse("finance:audit_log"),
            {"q": "journal", "action": FinanceAuditLog.Action.POSTED},
        )

        self.assertContains(response, "q=journal")
        self.assertContains(response, "action=posted")
        self.assertContains(response, "export=xlsx")
        self.assertContains(response, "export=pdf")


class ReportExportViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="report-export-accountant",
            password="test-password",
            role="accountant",
        )
        self.client.force_login(self.user)
        self.account = Account.objects.create(
            code="EX1100",
            name="Export cash",
            account_type=Account.AccountType.ASSET,
        )
        self.customer = Customer.objects.create(
            code="EX-C001",
            name="Export customer",
        )
        self.supplier = Supplier.objects.create(
            code="EX-S001",
            name="Export supplier",
        )

    def _excel_urls(self):
        return (
            f'{reverse("finance:general_ledger")}?account_id={self.account.pk}&export=xlsx',
            f'{reverse("finance:trial_balance")}?export=xlsx',
            f'{reverse("finance:income_statement")}?export=xlsx',
            f'{reverse("finance:balance_sheet")}?export=xlsx',
            f'{reverse("finance:cash_flow_statement")}?export=xlsx',
            f'{reverse("finance:customer_statement")}?customer={self.customer.pk}&export=xlsx',
            f'{reverse("finance:supplier_statement")}?supplier={self.supplier.pk}&export=xlsx',
            f'{reverse("finance:receivables_aging")}?export=xlsx',
            f'{reverse("finance:payables_aging")}?export=xlsx',
        )

    def test_every_finance_report_exports_a_valid_excel_workbook(self):
        for url in self._excel_urls():
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response["Content-Type"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.assertIn("attachment;", response["Content-Disposition"])
                self.assertTrue(response.content.startswith(b"PK"))
                workbook = load_workbook(BytesIO(response.content), read_only=True)
                self.assertTrue(workbook.active.title)

    def test_pdf_export_returns_a_pdf_download(self):
        response = self.client.get(
            reverse("finance:trial_balance"),
            {"export": "pdf"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("trial-balance.pdf", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertEqual(
            len(re.findall(rb"/Type\s*/Page\b", response.content)),
            1,
        )

    def test_arabic_excel_uses_rtl_sheet_direction(self):
        self.client.cookies[settings.LANGUAGE_COOKIE_NAME] = "ar"
        response = self.client.get(
            reverse("finance:trial_balance"),
            {"export": "xlsx"},
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertTrue(workbook.active.sheet_view.rightToLeft)
        self.assertEqual(workbook.active["A1"].value, "ميزان المراجعة")

    def test_report_page_shows_both_export_buttons(self):
        response = self.client.get(reverse("finance:trial_balance"))

        self.assertContains(response, "export=xlsx")
        self.assertContains(response, "export=pdf")

    @patch("finance.exports.pdf_response")
    def test_missing_pdf_runtime_returns_report_with_clear_error_toast(self, pdf_mock):
        pdf_mock.side_effect = PlaywrightError("Chromium executable not found")

        response = self.client.get(
            reverse("finance:trial_balance"),
            {"export": "pdf"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "PDF export is unavailable. Install the Playwright Chromium runtime",
        )


class ManualJournalViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="manual-accountant",
            password="test-password",
            role="accountant",
        )
        self.cash = Account.objects.create(
            code="M1100",
            name="Manual cash",
            account_type=Account.AccountType.ASSET,
        )
        self.revenue = Account.objects.create(
            code="M4100",
            name="Manual revenue",
            account_type=Account.AccountType.REVENUE,
        )
        self.customer = Customer.objects.create(code="MC001", name="Manual customer")
        self.supplier = Supplier.objects.create(code="MS001", name="Manual supplier")
        self.client.force_login(self.user)

    def journal_data(self, debit="100.000", credit="100.000"):
        return {
            "entry_number": "MJE-001",
            "date": date.today().isoformat(),
            "description": "Manual journal test",
            "cash_flow_activity": JournalEntry.CashFlowActivity.NONE,
            "lines-TOTAL_FORMS": "2",
            "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0",
            "lines-MAX_NUM_FORMS": "1000",
            "lines-0-account": str(self.cash.pk),
            "lines-0-customer": "",
            "lines-0-supplier": "",
            "lines-0-description": "Cash debit",
            "lines-0-debit": debit,
            "lines-0-credit": "0.000",
            "lines-1-account": str(self.revenue.pk),
            "lines-1-customer": "",
            "lines-1-supplier": "",
            "lines-1-description": "Revenue credit",
            "lines-1-debit": "0.000",
            "lines-1-credit": credit,
        }

    def test_balanced_manual_journal_is_saved_as_draft(self):
        response = self.client.post(reverse("finance:journal_create"), self.journal_data())

        journal = JournalEntry.objects.get(entry_number="MJE-001")
        self.assertRedirects(response, reverse("finance:journal_detail", args=[journal.pk]))
        self.assertEqual(journal.status, JournalEntry.Status.DRAFT)
        self.assertEqual(journal.source_type, JournalEntry.SourceType.MANUAL)
        self.assertEqual(journal.created_by, self.user)
        self.assertEqual(journal.lines.count(), 2)

    def test_unbalanced_manual_journal_is_rejected(self):
        response = self.client.post(
            reverse("finance:journal_create"),
            self.journal_data(credit="90.000"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Journal is not balanced")
        self.assertFalse(JournalEntry.objects.filter(entry_number="MJE-001").exists())

    def test_line_cannot_reference_customer_and_supplier(self):
        data = self.journal_data()
        data["lines-0-customer"] = str(self.customer.pk)
        data["lines-0-supplier"] = str(self.supplier.pk)

        response = self.client.post(reverse("finance:journal_create"), data)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "cannot reference both a customer and a supplier")
        self.assertFalse(JournalEntry.objects.filter(entry_number="MJE-001").exists())

    def test_create_page_contains_dynamic_line_controls(self):
        response = self.client.get(reverse("finance:journal_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="add-journal-line"')
        self.assertContains(response, 'id="journal-empty-line"')

    def test_draft_posting_error_is_shown_and_can_be_corrected(self):
        create_fiscal_year(date.today().year, user=self.user)
        self.cash.is_cash_equivalent = True
        self.cash.save(update_fields=["is_cash_equivalent"])
        self.client.post(reverse("finance:journal_create"), self.journal_data())
        journal = JournalEntry.objects.get(entry_number="MJE-001")

        detail = self.client.get(reverse("finance:journal_detail", args=[journal.pk]))

        self.assertContains(detail, "This journal is not ready to post")
        self.assertContains(detail, "Select a cash-flow activity")
        self.assertContains(detail, reverse("finance:journal_update", args=[journal.pk]))
        self.assertNotContains(detail, 'type="submit"><i class="bi bi-check2-circle"></i>Post entry')

        lines = list(journal.lines.order_by("pk"))
        update_data = self.journal_data()
        update_data.update(
            {
                "description": "Corrected manual journal",
                "cash_flow_activity": JournalEntry.CashFlowActivity.OPERATING,
                "lines-INITIAL_FORMS": "2",
                "lines-0-id": str(lines[0].pk),
                "lines-1-id": str(lines[1].pk),
            }
        )
        response = self.client.post(
            reverse("finance:journal_update", args=[journal.pk]),
            update_data,
        )

        self.assertRedirects(response, reverse("finance:journal_detail", args=[journal.pk]))
        journal.refresh_from_db()
        self.assertEqual(journal.description, "Corrected manual journal")
        self.assertEqual(journal.cash_flow_activity, JournalEntry.CashFlowActivity.OPERATING)
        self.assertTrue(
            FinanceAuditLog.objects.filter(
                action=FinanceAuditLog.Action.UPDATED,
                entity_type="finance.journalentry",
                object_id=str(journal.pk),
            ).exists()
        )

        corrected_detail = self.client.get(reverse("finance:journal_detail", args=[journal.pk]))
        self.assertNotContains(corrected_detail, "This journal is not ready to post")
        self.assertContains(corrected_detail, "Post entry")

    def test_posted_journal_cannot_be_edited(self):
        journal = JournalEntry.objects.create(
            entry_number="MJE-LOCKED",
            date=date.today(),
            source_type=JournalEntry.SourceType.MANUAL,
            status=JournalEntry.Status.POSTED,
            created_by=self.user,
        )

        response = self.client.get(
            reverse("finance:journal_update", args=[journal.pk]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("finance:journal_detail", args=[journal.pk]),
        )
        self.assertContains(response, "Only draft manual journal entries can be edited.")


class JournalDetailTransactionTests(TransactionTestCase):
    def test_draft_detail_uses_read_only_period_validation(self):
        user = get_user_model().objects.create_user(
            username="journal-detail-accountant",
            password="test-password",
            role="accountant",
        )
        journal = JournalEntry.objects.create(
            entry_number="MJE-READ-ONLY",
            date=date.today(),
            source_type=JournalEntry.SourceType.MANUAL,
            status=JournalEntry.Status.DRAFT,
            created_by=user,
        )
        self.client.force_login(user)

        response = self.client.get(reverse("finance:journal_detail", args=[journal.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This journal is not ready to post")
        self.assertContains(response, "No fiscal year has been configured")
        self.assertContains(response, reverse("finance:journal_update", args=[journal.pk]))


class FinancePostingTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="accountant",
            password="test-password",
            role="accountant",
        )
        create_required_fiscal_years(2026, date.today().year)
        self.customer = Customer.objects.create(code="C001", name="Customer")
        self.supplier = Supplier.objects.create(code="S001", name="Supplier")

        account_data = {
            "1100": ("Cash", Account.AccountType.ASSET),
            "1200": ("Bank", Account.AccountType.ASSET),
            "1210": ("Card clearing", Account.AccountType.ASSET),
            "1300": ("Accounts receivable", Account.AccountType.ASSET),
            "1400": ("Inventory", Account.AccountType.ASSET),
            "2100": ("Accounts payable", Account.AccountType.LIABILITY),
            "2200": ("Tax payable", Account.AccountType.LIABILITY),
            "4100": ("Sales revenue", Account.AccountType.REVENUE),
            "5100": ("Cost of goods sold", Account.AccountType.EXPENSE),
            "6300": ("Waste expense", Account.AccountType.EXPENSE),
        }
        self.accounts = {
            code: Account.objects.create(
                code=code,
                name=name,
                account_type=account_type,
                is_cash_equivalent=code in {"1100", "1200", "1210"},
            )
            for code, (name, account_type) in account_data.items()
        }

    def close_period_for(self, posting_date):
        period = FiscalPeriod.objects.get(
            fiscal_year__year=posting_date.year,
            month=posting_date.month,
        )
        set_period_status(
            period.pk,
            PeriodStatus.CLOSED,
            self.user,
            reason="Posting workflow validation test",
        )
        return period

    def test_post_journal_entry_records_approver(self):
        journal = JournalEntry.objects.create(
            entry_number="JE-001",
            date=date.today(),
            cash_flow_activity=JournalEntry.CashFlowActivity.OPERATING,
            created_by=self.user,
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.accounts["1100"],
            debit=Decimal("25.000"),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.accounts["4100"],
            credit=Decimal("25.000"),
        )

        post_journal_entry(journal.id, self.user)

        journal.refresh_from_db()
        self.assertEqual(journal.status, JournalEntry.Status.POSTED)
        self.assertEqual(journal.approved_by, self.user)

    def test_get_posting_account_rejects_wrong_account_type(self):
        revenue_account = self.accounts["4100"]
        revenue_account.account_type = Account.AccountType.EXPENSE
        revenue_account.save(update_fields=["account_type"])

        with self.assertRaisesMessage(ValidationError, "has the wrong type"):
            get_posting_account(
                "4100",
                Account.AccountType.REVENUE,
            )

    def test_unbalanced_journal_remains_draft(self):
        journal = JournalEntry.objects.create(
            entry_number="JE-002",
            date=date.today(),
            created_by=self.user,
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.accounts["1100"],
            debit=Decimal("25.000"),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.accounts["4100"],
            credit=Decimal("20.000"),
        )

        with self.assertRaises(ValidationError):
            post_journal_entry(journal.id, self.user)

        journal.refresh_from_db()
        self.assertEqual(journal.status, JournalEntry.Status.DRAFT)

    def test_receipt_and_payment_vouchers_post_balanced_journals(self):
        receipt = ReceiptVoucher.objects.create(
            voucher_number="001",
            date=date.today(),
            customer=self.customer,
            received_from=self.customer.name,
            account=self.accounts["1100"],
            amount=Decimal("40.000"),
            payment_method=ReceiptVoucher.PaymentMethod.CASH,
            created_by=self.user,
        )
        payment = PaymentVoucher.objects.create(
            voucher_number="001",
            date=date.today(),
            supplier=self.supplier,
            paid_to=self.supplier.name,
            account=self.accounts["1100"],
            amount=Decimal("30.000"),
            payment_method=PaymentVoucher.PaymentMethod.CASH,
            created_by=self.user,
        )

        post_receipt_voucher(receipt.id, self.user)
        post_payment_voucher(payment.id, self.user)

        receipt.refresh_from_db()
        payment.refresh_from_db()
        self.assertEqual(receipt.status, ReceiptVoucher.Status.CONFIRMED)
        self.assertEqual(payment.status, PaymentVoucher.Status.CONFIRMED)
        self.assertEqual(
            JournalEntry.objects.filter(status=JournalEntry.Status.POSTED).count(),
            2,
        )
        self.assertFalse(
            JournalEntry.objects.filter(
                cash_flow_activity=JournalEntry.CashFlowActivity.NONE,
            ).exists()
        )

    def test_source_can_only_be_posted_once(self):
        JournalEntry.objects.create(
            entry_number="SOURCE-1",
            date=date.today(),
            source_type=JournalEntry.SourceType.RECEIPT,
            source_id=99,
            created_by=self.user,
        )

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                JournalEntry.objects.create(
                    entry_number="SOURCE-2",
                    date=date.today(),
                    source_type=JournalEntry.SourceType.RECEIPT,
                    source_id=99,
                    created_by=self.user,
                )

    def test_pos_posts_split_revenue_tax_and_cogs(self):
        category = Category.objects.create(
            code="CAT", name_en="Category", name_ar="Category"
        )
        unit = Unit.objects.create(name_en="Piece", name_ar="Piece", symbol="pc")
        warehouse = Warehouse.objects.create(code="WH", name="Warehouse")
        product = Product.objects.create(
            code="P001",
            name_en="Product",
            name_ar="Product",
            category=category,
            unit=unit,
            purchase_price=Decimal("60.000"),
            selling_price=Decimal("100.000"),
        )
        sale = POSSale.objects.create(
            sale_number="SALE-001",
            warehouse=warehouse,
            cashier=self.user,
            status=POSSale.SaleStatus.COMPLETED,
            subtotal=Decimal("115.000"),
            total=Decimal("115.000"),
            paid_amount=Decimal("115.000"),
        )
        POSSaleItem.objects.create(
            sale=sale,
            product=product,
            quantity=Decimal("1.000"),
            unit_price=Decimal("100.000"),
            unit_cost=Decimal("60.000"),
            tax_amount=Decimal("15.000"),
            line_total=Decimal("115.000"),
        )
        POSPayment.objects.create(
            sale=sale,
            payment_method=POSPayment.PaymentMethod.CASH,
            amount=Decimal("115.000"),
            created_by=self.user,
        )

        journal = post_pos_sale(sale.id, self.user)

        amounts = {
            line.account.code: (line.debit, line.credit)
            for line in journal.lines.select_related("account")
        }
        self.assertEqual(amounts["1100"], (Decimal("115.000"), Decimal("0.000")))
        self.assertEqual(amounts["4100"], (Decimal("0.000"), Decimal("100.000")))
        self.assertEqual(amounts["2200"], (Decimal("0.000"), Decimal("15.000")))
        self.assertEqual(amounts["5100"], (Decimal("60.000"), Decimal("0.000")))
        self.assertEqual(amounts["1400"], (Decimal("0.000"), Decimal("60.000")))

        with self.assertRaises(ValidationError):
            post_pos_sale(sale.id, self.user)

    def test_purchase_uses_one_balanced_finance_journal(self):
        category = Category.objects.create(
            code="PCAT", name_en="Purchase Category", name_ar="Purchase Category"
        )
        unit = Unit.objects.create(name_en="Box", name_ar="Box", symbol="box")
        warehouse = Warehouse.objects.create(code="PWH", name="Purchase Warehouse")
        product = Product.objects.create(
            code="PP01",
            name_en="Purchased product",
            name_ar="Purchased product",
            category=category,
            unit=unit,
            purchase_price=Decimal("50.000"),
            selling_price=Decimal("75.000"),
        )
        invoice = PurchaseInvoice.objects.create(
            invoice_number="PI-001",
            supplier=self.supplier,
            warehouse=warehouse,
            invoice_date=date.today(),
            payment_type=PurchaseInvoice.PaymentType.CREDIT,
            status=PurchaseInvoice.Status.CONFIRMED,
            subtotal=Decimal("100.000"),
            total=Decimal("100.000"),
            created_by=self.user,
        )
        PurchaseInvoiceItem.objects.create(
            purchase_invoice=invoice,
            product=product,
            quantity=Decimal("2.000"),
            unit_cost=Decimal("50.000"),
            line_total=Decimal("100.000"),
        )

        journal = post_purchase_invoice(invoice.id, self.user)

        self.assertEqual(journal.status, JournalEntry.Status.POSTED)
        self.assertEqual(journal.lines.count(), 2)
        self.assertEqual(
            sum((line.debit for line in journal.lines.all()), Decimal("0.000")),
            sum((line.credit for line in journal.lines.all()), Decimal("0.000")),
        )

    def test_complete_sale_posts_finance_with_actual_batch_cost(self):
        category = Category.objects.create(
            code="SCAT", name_en="Sale Category", name_ar="Sale Category"
        )
        unit = Unit.objects.create(name_en="Unit", name_ar="Unit", symbol="u")
        warehouse = Warehouse.objects.create(code="SWH", name="Sale Warehouse")
        product = Product.objects.create(
            code="SP01",
            name_en="Stock product",
            name_ar="Stock product",
            category=category,
            unit=unit,
            purchase_price=Decimal("30.000"),
            selling_price=Decimal("80.000"),
        )
        StockBatch.objects.create(
            product=product,
            warehouse=warehouse,
            batch_number="B001",
            received_date=date.today(),
            unit_cost=Decimal("45.000"),
            quantity_received=Decimal("5.000"),
            quantity_remaining=Decimal("5.000"),
        )
        StockBalance.objects.create(
            product=product,
            warehouse=warehouse,
            quantity=Decimal("5.000"),
        )

        sale = complete_sale(
            warehouse=warehouse,
            cashier=self.user,
            items_data=[{
                "product": product,
                "quantity": "2.000",
                "unit_price": "80.000",
            }],
            payments_data=[{
                "payment_method": POSPayment.PaymentMethod.CASH,
                "amount": "160.000",
            }],
        )

        sale_item = sale.items.get()
        journal = JournalEntry.objects.get(
            source_type=JournalEntry.SourceType.POS_SALE,
            source_id=sale.id,
        )
        self.assertEqual(sale_item.unit_cost, Decimal("45.000"))
        self.assertEqual(journal.status, JournalEntry.Status.POSTED)
        self.assertEqual(
            journal.lines.get(account=self.accounts["5100"]).debit,
            Decimal("90.000"),
        )

    def test_confirm_waste_loss_posts_finance_journal(self):
        category = Category.objects.create(
            code="WCAT", name_en="Waste Category", name_ar="Waste Category"
        )
        unit = Unit.objects.create(
            name_en="Waste Unit", name_ar="Waste Unit", symbol="wu"
        )
        warehouse = Warehouse.objects.create(code="WWH", name="Waste Warehouse")
        product = Product.objects.create(
            code="WP01",
            name_en="Waste product",
            name_ar="Waste product",
            category=category,
            unit=unit,
            purchase_price=Decimal("12.000"),
            selling_price=Decimal("20.000"),
        )
        StockBatch.objects.create(
            product=product,
            warehouse=warehouse,
            batch_number="WB001",
            received_date=date.today(),
            unit_cost=Decimal("12.000"),
            quantity_received=Decimal("2.000"),
            quantity_remaining=Decimal("2.000"),
        )
        StockBalance.objects.create(
            product=product,
            warehouse=warehouse,
            quantity=Decimal("2.000"),
        )
        waste = WasteLoss.objects.create(
            document_number="WST-001",
            warehouse=warehouse,
            date=date.today(),
            reason=WasteLoss.WasteReason.DAMAGED,
            created_by=self.user,
        )
        WasteLossItem.objects.create(
            waste_loss=waste,
            product=product,
            quantity=Decimal("2.000"),
            unit_cost=Decimal("12.000"),
            total_cost=Decimal("24.000"),
        )

        confirmed_waste, journal = confirm_waste_loss(waste.id, self.user)

        self.assertEqual(confirmed_waste.status, WasteLoss.Status.CONFIRMED)
        self.assertEqual(journal.status, JournalEntry.Status.POSTED)
        self.assertEqual(
            journal.lines.get(account=self.accounts["6300"]).debit,
            Decimal("24.000"),
        )
        self.assertEqual(
            journal.lines.get(account=self.accounts["1400"]).credit,
            Decimal("24.000"),
        )

    def test_waste_confirmation_rolls_back_when_finance_posting_fails(self):
        category = Category.objects.create(
            code="RCAT", name_en="Rollback Category", name_ar="Rollback Category"
        )
        unit = Unit.objects.create(
            name_en="Rollback Unit", name_ar="Rollback Unit", symbol="ru"
        )
        warehouse = Warehouse.objects.create(code="RWH", name="Rollback Warehouse")
        product = Product.objects.create(
            code="RP01",
            name_en="Rollback product",
            name_ar="Rollback product",
            category=category,
            unit=unit,
            purchase_price=Decimal("10.000"),
            selling_price=Decimal("15.000"),
        )
        waste = WasteLoss.objects.create(
            document_number="WST-ROLLBACK",
            warehouse=warehouse,
            date=date.today(),
            reason=WasteLoss.WasteReason.DAMAGED,
            created_by=self.user,
        )
        WasteLossItem.objects.create(
            waste_loss=waste,
            product=product,
            quantity=Decimal("1.000"),
            unit_cost=Decimal("10.000"),
            total_cost=Decimal("10.000"),
        )
        self.accounts["6300"].is_active = False
        self.accounts["6300"].save(update_fields=["is_active"])

        with self.assertRaises(ValidationError):
            confirm_waste_loss(waste.id, self.user)

        waste.refresh_from_db()
        self.assertEqual(waste.status, WasteLoss.Status.DRAFT)
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.WASTE,
                source_id=waste.id,
            ).exists()
        )

    def test_closed_period_rejects_receipt_and_rolls_back_journal(self):
        self.close_period_for(date.today())
        receipt = ReceiptVoucher.objects.create(
            voucher_number="CLOSED-RV",
            date=date.today(),
            customer=self.customer,
            received_from=self.customer.name,
            account=self.accounts["1100"],
            amount=Decimal("40.000"),
            payment_method=ReceiptVoucher.PaymentMethod.CASH,
            created_by=self.user,
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            post_receipt_voucher(receipt.pk, self.user)

        receipt.refresh_from_db()
        self.assertEqual(receipt.status, ReceiptVoucher.Status.DRAFT)
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.RECEIPT,
                source_id=receipt.pk,
            ).exists()
        )

    def test_closed_period_rejects_payment_and_rolls_back_journal(self):
        self.close_period_for(date.today())
        payment = PaymentVoucher.objects.create(
            voucher_number="CLOSED-PV",
            date=date.today(),
            supplier=self.supplier,
            paid_to=self.supplier.name,
            account=self.accounts["1100"],
            amount=Decimal("30.000"),
            payment_method=PaymentVoucher.PaymentMethod.CASH,
            created_by=self.user,
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            post_payment_voucher(payment.pk, self.user)

        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentVoucher.Status.DRAFT)
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.PAYMENT,
                source_id=payment.pk,
            ).exists()
        )

    def test_closed_period_rejects_purchase_and_rolls_back_inventory(self):
        self.close_period_for(date.today())
        category = Category.objects.create(
            code="CLOSED-PCAT",
            name_en="Closed purchase category",
            name_ar="Closed purchase category",
        )
        unit = Unit.objects.create(
            name_en="Closed purchase unit",
            name_ar="Closed purchase unit",
            symbol="cpu",
        )
        warehouse = Warehouse.objects.create(
            code="CLOSED-PWH",
            name="Closed purchase warehouse",
        )
        product = Product.objects.create(
            code="CLOSED-PP",
            name_en="Closed purchase product",
            name_ar="Closed purchase product",
            category=category,
            unit=unit,
            purchase_price=Decimal("25.000"),
            selling_price=Decimal("40.000"),
        )
        invoice = PurchaseInvoice.objects.create(
            invoice_number="CLOSED-PI",
            supplier=self.supplier,
            warehouse=warehouse,
            invoice_date=date.today(),
            payment_type=PurchaseInvoice.PaymentType.CREDIT,
            subtotal=Decimal("50.000"),
            total=Decimal("50.000"),
            created_by=self.user,
        )
        PurchaseInvoiceItem.objects.create(
            purchase_invoice=invoice,
            product=product,
            quantity=Decimal("2.000"),
            unit_cost=Decimal("25.000"),
            line_total=Decimal("50.000"),
            batch_number="CLOSED-BATCH",
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            confirm_purchase(invoice.pk, self.user)

        invoice.refresh_from_db()
        self.assertEqual(invoice.status, PurchaseInvoice.Status.DRAFT)
        self.assertFalse(StockBatch.objects.filter(product=product).exists())
        self.assertFalse(StockBalance.objects.filter(product=product).exists())
        self.assertFalse(StockMovement.objects.filter(product=product).exists())
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.PURCHASE,
                source_id=invoice.pk,
            ).exists()
        )

    def test_closed_period_rejects_pos_sale_and_restores_stock(self):
        self.close_period_for(date.today())
        category = Category.objects.create(
            code="CLOSED-SCAT",
            name_en="Closed sale category",
            name_ar="Closed sale category",
        )
        unit = Unit.objects.create(
            name_en="Closed sale unit",
            name_ar="Closed sale unit",
            symbol="csu",
        )
        warehouse = Warehouse.objects.create(
            code="CLOSED-SWH",
            name="Closed sale warehouse",
        )
        product = Product.objects.create(
            code="CLOSED-SP",
            name_en="Closed sale product",
            name_ar="Closed sale product",
            category=category,
            unit=unit,
            purchase_price=Decimal("20.000"),
            selling_price=Decimal("35.000"),
        )
        batch = StockBatch.objects.create(
            product=product,
            warehouse=warehouse,
            batch_number="CLOSED-SALE-BATCH",
            received_date=date.today(),
            unit_cost=Decimal("20.000"),
            quantity_received=Decimal("5.000"),
            quantity_remaining=Decimal("5.000"),
        )
        balance = StockBalance.objects.create(
            product=product,
            warehouse=warehouse,
            quantity=Decimal("5.000"),
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            complete_sale(
                warehouse=warehouse,
                cashier=self.user,
                items_data=[{
                    "product": product,
                    "quantity": "2.000",
                    "unit_price": "35.000",
                }],
                payments_data=[{
                    "payment_method": POSPayment.PaymentMethod.CASH,
                    "amount": "70.000",
                }],
            )

        batch.refresh_from_db()
        balance.refresh_from_db()
        self.assertEqual(batch.quantity_remaining, Decimal("5.000"))
        self.assertEqual(balance.quantity, Decimal("5.000"))
        self.assertFalse(POSSale.objects.filter(warehouse=warehouse).exists())
        self.assertFalse(StockMovement.objects.filter(product=product).exists())
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.POS_SALE,
            ).exists()
        )

    def test_closed_period_rejects_waste_and_restores_stock(self):
        self.close_period_for(date.today())
        category = Category.objects.create(
            code="CLOSED-WCAT",
            name_en="Closed waste category",
            name_ar="Closed waste category",
        )
        unit = Unit.objects.create(
            name_en="Closed waste unit",
            name_ar="Closed waste unit",
            symbol="cwu",
        )
        warehouse = Warehouse.objects.create(
            code="CLOSED-WWH",
            name="Closed waste warehouse",
        )
        product = Product.objects.create(
            code="CLOSED-WP",
            name_en="Closed waste product",
            name_ar="Closed waste product",
            category=category,
            unit=unit,
            purchase_price=Decimal("12.000"),
            selling_price=Decimal("18.000"),
        )
        batch = StockBatch.objects.create(
            product=product,
            warehouse=warehouse,
            batch_number="CLOSED-WASTE-BATCH",
            received_date=date.today(),
            unit_cost=Decimal("12.000"),
            quantity_received=Decimal("3.000"),
            quantity_remaining=Decimal("3.000"),
        )
        balance = StockBalance.objects.create(
            product=product,
            warehouse=warehouse,
            quantity=Decimal("3.000"),
        )
        waste = WasteLoss.objects.create(
            document_number="CLOSED-WASTE",
            warehouse=warehouse,
            date=date.today(),
            reason=WasteLoss.WasteReason.DAMAGED,
            created_by=self.user,
        )
        WasteLossItem.objects.create(
            waste_loss=waste,
            product=product,
            batch=batch,
            quantity=Decimal("1.000"),
            unit_cost=Decimal("12.000"),
            total_cost=Decimal("12.000"),
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            confirm_waste_loss(waste.pk, self.user)

        waste.refresh_from_db()
        batch.refresh_from_db()
        balance.refresh_from_db()
        self.assertEqual(waste.status, WasteLoss.Status.DRAFT)
        self.assertEqual(batch.quantity_remaining, Decimal("3.000"))
        self.assertEqual(balance.quantity, Decimal("3.000"))
        self.assertFalse(StockMovement.objects.filter(product=product).exists())
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.WASTE,
                source_id=waste.pk,
            ).exists()
        )


class JournalReversalTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="reversal-accountant",
            password="test-password",
            role="accountant",
        )
        create_required_fiscal_years(date.today().year)
        self.cash = Account.objects.create(
            code="REV1100",
            name="Reversal cash",
            account_type=Account.AccountType.ASSET,
        )
        self.revenue = Account.objects.create(
            code="REV4100",
            name="Reversal revenue",
            account_type=Account.AccountType.REVENUE,
        )

    def create_journal(self, *, status=JournalEntry.Status.POSTED, source_type=None):
        journal = JournalEntry.objects.create(
            entry_number=f"REV-TEST-{JournalEntry.objects.count() + 1}",
            date=date.today(),
            description="Journal to reverse",
            source_type=source_type or JournalEntry.SourceType.MANUAL,
            status=status,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.cash,
            description="Cash debit",
            debit=Decimal("75.000"),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.revenue,
            description="Revenue credit",
            credit=Decimal("75.000"),
        )
        return journal

    def test_reversal_posts_opposite_lines_and_marks_original_reversed(self):
        original = self.create_journal()

        reversal = reverse_journal_entry(
            original.pk,
            self.accountant,
            "Incorrect customer",
        )

        original.refresh_from_db()
        self.assertEqual(original.status, JournalEntry.Status.REVERSED)
        self.assertEqual(original.reversal_reason, "Incorrect customer")
        self.assertEqual(original.reversed_by, self.accountant)
        self.assertIsNotNone(original.reversed_at)

        self.assertEqual(reversal.status, JournalEntry.Status.POSTED)
        self.assertEqual(reversal.source_type, JournalEntry.SourceType.REVERSAL)
        self.assertEqual(reversal.source_id, original.pk)
        self.assertEqual(reversal.reversal_of, original)
        self.assertEqual(reversal.approved_by, self.accountant)
        self.assertEqual(
            reversal.lines.get(account=self.cash).credit,
            Decimal("75.000"),
        )
        self.assertEqual(
            reversal.lines.get(account=self.revenue).debit,
            Decimal("75.000"),
        )
        self.assertEqual(get_account_balance(self.cash), Decimal("0.000"))
        self.assertEqual(get_account_balance(self.revenue), Decimal("0.000"))

    def test_journal_cannot_be_reversed_twice(self):
        original = self.create_journal()
        reverse_journal_entry(original.pk, self.accountant, "First reversal")

        with self.assertRaisesMessage(
            ValidationError,
            "This journal entry has already been reversed.",
        ):
            reverse_journal_entry(original.pk, self.accountant, "Again")

        self.assertEqual(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.REVERSAL,
                source_id=original.pk,
            ).count(),
            1,
        )

    def test_reversal_requires_posted_entry_and_reason(self):
        draft = self.create_journal(status=JournalEntry.Status.DRAFT)
        with self.assertRaisesMessage(
            ValidationError,
            "Only posted journal entries can be reversed.",
        ):
            reverse_journal_entry(draft.pk, self.accountant, "Correction")

        posted = self.create_journal()
        with self.assertRaisesMessage(
            ValidationError,
            "A reversal reason is required.",
        ):
            reverse_journal_entry(posted.pk, self.accountant, None)

    def test_reversal_entry_cannot_be_reversed(self):
        reversal = self.create_journal(
            source_type=JournalEntry.SourceType.REVERSAL,
        )

        with self.assertRaisesMessage(
            ValidationError,
            "A reversal journal entry cannot be reversed.",
        ):
            reverse_journal_entry(reversal.pk, self.accountant, "Undo")

    def test_reversal_view_requires_post_and_accountant_permission(self):
        original = self.create_journal()
        url = reverse("finance:journal_reverse", args=[original.pk])
        self.client.force_login(self.accountant)
        self.assertEqual(self.client.get(url).status_code, 405)

        cashier = get_user_model().objects.create_user(
            username="reversal-cashier",
            password="test-password",
            role="cashier",
        )
        self.client.force_login(cashier)
        self.assertEqual(
            self.client.post(url, {"reason": "Not allowed"}).status_code,
            403,
        )
        original.refresh_from_db()
        self.assertEqual(original.status, JournalEntry.Status.POSTED)

    def test_reversal_view_redirects_to_created_reversal(self):
        original = self.create_journal()
        self.client.force_login(self.accountant)

        response = self.client.post(
            reverse("finance:journal_reverse", args=[original.pk]),
            {"reason": "Entered twice"},
        )

        reversal = JournalEntry.objects.get(reversal_of=original)
        self.assertRedirects(
            response,
            reverse("finance:journal_detail", args=[reversal.pk]),
        )
        original_response = self.client.get(
            reverse("finance:journal_detail", args=[original.pk])
        )
        self.assertContains(original_response, "Entered twice")
        self.assertContains(original_response, reversal.entry_number)

    def test_closed_period_rejects_reversal_and_preserves_original(self):
        original = self.create_journal()
        period = FiscalPeriod.objects.get(
            fiscal_year__year=date.today().year,
            month=date.today().month,
        )
        set_period_status(
            period.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Month-end close",
        )

        with self.assertRaisesMessage(ValidationError, "period"):
            reverse_journal_entry(
                original.pk,
                self.accountant,
                "Correction after close",
            )

        original.refresh_from_db()
        self.assertEqual(original.status, JournalEntry.Status.POSTED)
        self.assertFalse(
            JournalEntry.objects.filter(
                source_type=JournalEntry.SourceType.REVERSAL,
                source_id=original.pk,
            ).exists()
        )


class IncomeStatementTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="statement-accountant",
            password="test-password",
            role="accountant",
        )
        create_required_fiscal_years(2026, date.today().year)
        self.cash = Account.objects.create(
            code="IS1100",
            name="Statement cash",
            account_type=Account.AccountType.ASSET,
        )
        self.revenue = Account.objects.create(
            code="IS4100",
            name="Statement revenue",
            account_type=Account.AccountType.REVENUE,
        )
        self.expense = Account.objects.create(
            code="IS5100",
            name="Statement expense",
            account_type=Account.AccountType.EXPENSE,
        )

    def create_journal(self, number, journal_date, lines, *, status):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=journal_date,
            description="Income statement test",
            status=status,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        for account, debit, credit in lines:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=account,
                debit=debit,
                credit=credit,
            )
        return journal

    def test_income_statement_calculates_profit_and_excludes_drafts(self):
        self.create_journal(
            "IS-SALE",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("100.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("100.000")),
            ],
            status=JournalEntry.Status.POSTED,
        )
        self.create_journal(
            "IS-EXPENSE",
            date(2026, 8, 2),
            [
                (self.expense, Decimal("40.000"), Decimal("0.000")),
                (self.cash, Decimal("0.000"), Decimal("40.000")),
            ],
            status=JournalEntry.Status.POSTED,
        )
        self.create_journal(
            "IS-DRAFT",
            date(2026, 8, 3),
            [
                (self.cash, Decimal("500.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("500.000")),
            ],
            status=JournalEntry.Status.DRAFT,
        )

        statement = generate_income_statement()

        self.assertEqual(statement["total_revenue"], Decimal("100.000"))
        self.assertEqual(statement["total_expenses"], Decimal("40.000"))
        self.assertEqual(statement["net_profit"], Decimal("60.000"))

    def test_income_statement_applies_date_range(self):
        self.create_journal(
            "IS-JULY",
            date(2026, 7, 31),
            [
                (self.cash, Decimal("80.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("80.000")),
            ],
            status=JournalEntry.Status.POSTED,
        )
        self.create_journal(
            "IS-AUGUST",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("25.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("25.000")),
            ],
            status=JournalEntry.Status.POSTED,
        )

        statement = generate_income_statement(
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )

        self.assertEqual(statement["total_revenue"], Decimal("25.000"))

    def test_reversal_cancels_income_statement_activity(self):
        original = self.create_journal(
            "IS-REVERSED",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("70.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("70.000")),
            ],
            status=JournalEntry.Status.POSTED,
        )
        reverse_journal_entry(original.pk, self.accountant, "Incorrect sale")

        statement = generate_income_statement()

        self.assertEqual(statement["total_revenue"], Decimal("0.000"))
        self.assertEqual(statement["net_profit"], Decimal("0.000"))
        self.assertEqual(statement["revenue_rows"], [])

    def test_income_statement_view_validates_dates(self):
        self.client.force_login(self.accountant)
        url = reverse("finance:income_statement")

        invalid_format = self.client.get(url, {"start_date": "not-a-date"})
        self.assertEqual(invalid_format.status_code, 200)
        self.assertContains(invalid_format, "Enter a valid date.")

        reversed_range = self.client.get(
            url,
            {"start_date": "2026-08-31", "end_date": "2026-08-01"},
        )
        self.assertEqual(reversed_range.status_code, 200)
        self.assertContains(
            reversed_range,
            "The start date cannot be later than the end date.",
        )


class BalanceSheetTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="balance-accountant",
            password="test-password",
            role="accountant",
        )
        create_required_fiscal_years(2026, date.today().year)
        self.cash = Account.objects.create(
            code="BS1100",
            name="Balance cash",
            account_type=Account.AccountType.ASSET,
        )
        self.payable = Account.objects.create(
            code="BS2100",
            name="Balance payable",
            account_type=Account.AccountType.LIABILITY,
        )
        self.capital = Account.objects.create(
            code="BS3100",
            name="Owner capital",
            account_type=Account.AccountType.EQUITY,
        )
        self.revenue = Account.objects.create(
            code="BS4100",
            name="Balance revenue",
            account_type=Account.AccountType.REVENUE,
        )
        self.expense = Account.objects.create(
            code="BS5100",
            name="Balance expense",
            account_type=Account.AccountType.EXPENSE,
        )

    def create_journal(self, number, journal_date, lines, *, status=JournalEntry.Status.POSTED):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=journal_date,
            description="Balance sheet test",
            status=status,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        for account, debit, credit in lines:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=account,
                debit=debit,
                credit=credit,
            )
        return journal

    def test_balance_sheet_includes_equity_liabilities_and_current_earnings(self):
        self.create_journal(
            "BS-CAPITAL",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("100.000"), Decimal("0.000")),
                (self.capital, Decimal("0.000"), Decimal("100.000")),
            ],
        )
        self.create_journal(
            "BS-PAYABLE",
            date(2026, 8, 2),
            [
                (self.cash, Decimal("30.000"), Decimal("0.000")),
                (self.payable, Decimal("0.000"), Decimal("30.000")),
            ],
        )
        self.create_journal(
            "BS-PROFIT",
            date(2026, 8, 3),
            [
                (self.cash, Decimal("20.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("30.000")),
                (self.expense, Decimal("10.000"), Decimal("0.000")),
            ],
        )

        report = generate_balance_sheet()

        self.assertEqual(report["total_assets"], Decimal("150.000"))
        self.assertEqual(report["total_liabilities"], Decimal("30.000"))
        self.assertEqual(report["total_equity"], Decimal("100.000"))
        self.assertEqual(report["current_earnings"], Decimal("20.000"))
        self.assertEqual(
            report["total_liabilities_and_equity"],
            Decimal("150.000"),
        )
        self.assertTrue(report["is_balanced"])

    def test_balance_sheet_excludes_drafts_and_entries_after_as_of_date(self):
        self.create_journal(
            "BS-IN-RANGE",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("40.000"), Decimal("0.000")),
                (self.capital, Decimal("0.000"), Decimal("40.000")),
            ],
        )
        self.create_journal(
            "BS-AFTER",
            date(2026, 9, 1),
            [
                (self.cash, Decimal("50.000"), Decimal("0.000")),
                (self.capital, Decimal("0.000"), Decimal("50.000")),
            ],
        )
        self.create_journal(
            "BS-DRAFT",
            date(2026, 8, 2),
            [
                (self.cash, Decimal("60.000"), Decimal("0.000")),
                (self.capital, Decimal("0.000"), Decimal("60.000")),
            ],
            status=JournalEntry.Status.DRAFT,
        )

        report = generate_balance_sheet(as_of_date=date(2026, 8, 31))

        self.assertEqual(report["total_assets"], Decimal("40.000"))
        self.assertEqual(report["total_equity"], Decimal("40.000"))
        self.assertTrue(report["is_balanced"])

    def test_reversal_cancels_balance_sheet_activity(self):
        original = self.create_journal(
            "BS-REVERSED",
            date(2026, 8, 1),
            [
                (self.cash, Decimal("70.000"), Decimal("0.000")),
                (self.revenue, Decimal("0.000"), Decimal("70.000")),
            ],
        )
        reverse_journal_entry(original.pk, self.accountant, "Incorrect sale")

        report = generate_balance_sheet()

        self.assertEqual(report["total_assets"], Decimal("0.000"))
        self.assertEqual(report["current_earnings"], Decimal("0.000"))
        self.assertTrue(report["is_balanced"])

    def test_balance_sheet_view_validates_date_and_restricts_cashiers(self):
        url = reverse("finance:balance_sheet")
        self.client.force_login(self.accountant)

        invalid = self.client.get(url, {"as_of_date": "not-a-date"})
        self.assertEqual(invalid.status_code, 200)
        self.assertContains(invalid, "Enter a valid date.")

        cashier = get_user_model().objects.create_user(
            username="balance-cashier",
            password="test-password",
            role="cashier",
        )
        self.client.force_login(cashier)
        self.assertEqual(self.client.get(url).status_code, 403)


class FiscalPeriodTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="period-accountant",
            password="test-password",
            role="accountant",
        )
        self.admin = get_user_model().objects.create_user(
            username="period-admin",
            password="test-password",
            role="admin",
        )
        self.cashier = get_user_model().objects.create_user(
            username="period-cashier",
            password="test-password",
            role="cashier",
        )
        self.fiscal_year = create_fiscal_year(2026)
        self.period = self.fiscal_year.periods.get(month=8)
        self.cash = Account.objects.create(
            code="FP1100",
            name="Period cash",
            account_type=Account.AccountType.ASSET,
        )
        self.capital = Account.objects.create(
            code="FP3100",
            name="Period capital",
            account_type=Account.AccountType.EQUITY,
        )

    def create_draft_journal(self, number="FP-JE-001"):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=date(2026, 8, 15),
            description="Fiscal period test",
            created_by=self.accountant,
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.cash,
            debit=Decimal("10.000"),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.capital,
            credit=Decimal("10.000"),
        )
        return journal

    def test_year_creation_builds_twelve_calendar_periods(self):
        periods = list(self.fiscal_year.periods.order_by("month"))

        self.assertEqual(len(periods), 12)
        self.assertEqual(periods[0].start_date, date(2026, 1, 1))
        self.assertEqual(periods[1].end_date, date(2026, 2, 28))
        self.assertTrue(
            all(period.status == PeriodStatus.OPEN for period in periods)
        )

    def test_closed_month_blocks_posting(self):
        set_period_status(
            self.period.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Monthly close",
        )
        journal = self.create_draft_journal()

        with self.assertRaisesMessage(ValidationError, "accounting period"):
            post_journal_entry(journal.pk, self.accountant)

        journal.refresh_from_db()
        self.assertEqual(journal.status, JournalEntry.Status.DRAFT)

    def test_reopening_month_allows_posting_and_clears_audit_fields(self):
        set_period_status(
            self.period.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Monthly close",
        )
        set_period_status(
            self.period.pk,
            PeriodStatus.OPEN,
            self.admin,
            reason="Approved correction",
        )
        journal = self.create_draft_journal()

        post_journal_entry(journal.pk, self.accountant)

        self.period.refresh_from_db()
        journal.refresh_from_db()
        self.assertIsNone(self.period.closed_by)
        self.assertIsNone(self.period.closed_at)
        self.assertEqual(journal.status, JournalEntry.Status.POSTED)

    def test_closing_year_closes_every_month(self):
        set_fiscal_year_status(
            self.fiscal_year.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Year-end review complete",
        )

        self.fiscal_year.refresh_from_db()
        self.assertEqual(self.fiscal_year.status, PeriodStatus.CLOSED)
        self.assertEqual(self.fiscal_year.closed_by, self.accountant)
        self.assertEqual(
            self.fiscal_year.periods.filter(status=PeriodStatus.CLOSED).count(),
            12,
        )

    def test_closed_year_blocks_posting(self):
        set_fiscal_year_status(
            self.fiscal_year.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Year-end review complete",
        )
        journal = self.create_draft_journal()

        with self.assertRaisesMessage(ValidationError, "fiscal year is closed"):
            post_journal_entry(journal.pk, self.accountant)

    def test_month_cannot_reopen_while_year_is_closed(self):
        set_fiscal_year_status(
            self.fiscal_year.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Year-end review complete",
        )

        with self.assertRaisesMessage(ValidationError, "Open the fiscal year"):
            set_period_status(
                self.period.pk,
                PeriodStatus.OPEN,
                self.admin,
                reason="Correction required",
            )

    def test_only_administrator_can_reopen_month_or_year(self):
        set_period_status(
            self.period.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Monthly close",
        )

        with self.assertRaises(PermissionDenied):
            set_period_status(
                self.period.pk,
                PeriodStatus.OPEN,
                self.accountant,
                reason="Correction",
            )

        set_fiscal_year_status(
            self.fiscal_year.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Year close",
        )
        with self.assertRaises(PermissionDenied):
            set_fiscal_year_status(
                self.fiscal_year.pk,
                PeriodStatus.OPEN,
                self.accountant,
                reason="Correction",
            )

        set_fiscal_year_status(
            self.fiscal_year.pk,
            PeriodStatus.OPEN,
            self.admin,
            reason="Approved correction",
        )
        self.fiscal_year.refresh_from_db()
        self.assertEqual(self.fiscal_year.status, PeriodStatus.OPEN)

    def test_accountant_can_manage_periods_but_cashier_cannot(self):
        list_url = reverse("finance:fiscal_period_list")
        status_url = reverse("finance:fiscal_period_status", args=[self.period.pk])

        self.client.force_login(self.accountant)
        self.assertEqual(self.client.get(list_url).status_code, 200)
        response = self.client.post(
            status_url,
            {"status": PeriodStatus.CLOSED, "reason": "Monthly close"},
        )
        self.assertRedirects(response, list_url)
        self.assertEqual(
            self.client.post(
                status_url,
                {"status": PeriodStatus.OPEN, "reason": "Correction"},
            ).status_code,
            403,
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            status_url,
            {"status": PeriodStatus.OPEN, "reason": "Approved correction"},
        )
        self.assertRedirects(response, list_url)

        self.client.force_login(self.cashier)
        self.assertEqual(self.client.get(list_url).status_code, 403)
        self.assertEqual(
            self.client.post(
                status_url,
                {"status": PeriodStatus.OPEN, "reason": "Not allowed"},
            ).status_code,
            403,
        )

    def test_reason_is_required_and_actions_are_audited(self):
        with self.assertRaisesMessage(ValidationError, "reason is required"):
            set_period_status(
                self.period.pk,
                PeriodStatus.CLOSED,
                self.accountant,
            )

        set_period_status(
            self.period.pk,
            PeriodStatus.CLOSED,
            self.accountant,
            reason="Books reconciled",
        )
        action = FiscalPeriodAction.objects.get(period=self.period)
        self.assertEqual(action.action, FiscalPeriodAction.Action.CLOSED)
        self.assertEqual(action.performed_by, self.accountant)
        self.assertEqual(action.reason, "Books reconciled")

    def test_draft_journal_blocks_closing(self):
        self.create_draft_journal()

        with self.assertRaisesMessage(ValidationError, "draft journals"):
            set_period_status(
                self.period.pk,
                PeriodStatus.CLOSED,
                self.accountant,
                reason="Attempted close",
            )

        self.period.refresh_from_db()
        self.assertEqual(self.period.status, PeriodStatus.OPEN)
        self.assertFalse(FiscalPeriodAction.objects.filter(period=self.period).exists())

    def test_period_summary_reports_posted_totals(self):
        journal = self.create_draft_journal()
        post_journal_entry(journal.pk, self.accountant)

        summary = get_period_summary(self.period)

        self.assertEqual(summary["posted_journals"], 1)
        self.assertEqual(summary["draft_journals"], 0)
        self.assertEqual(summary["total_debit"], Decimal("10.000"))
        self.assertEqual(summary["total_credit"], Decimal("10.000"))

    def test_accountant_can_update_period_notes(self):
        self.client.force_login(self.accountant)
        response = self.client.post(
            reverse("finance:fiscal_period_notes", args=[self.period.pk]),
            {"notes": "Bank reconciliation completed."},
        )

        self.assertRedirects(response, reverse("finance:fiscal_period_list"))
        self.period.refresh_from_db()
        self.assertEqual(self.period.notes, "Bank reconciliation completed.")

    def test_history_page_is_linked_and_paginates_ten_rows(self):
        FiscalPeriodAction.objects.bulk_create(
            [
                FiscalPeriodAction(
                    fiscal_year=self.fiscal_year,
                    period=self.period,
                    action=(
                        FiscalPeriodAction.Action.CLOSED
                        if index % 2
                        else FiscalPeriodAction.Action.OPENED
                    ),
                    performed_by=self.admin,
                    reason=f"History action {index}",
                )
                for index in range(11)
            ]
        )
        list_url = reverse("finance:fiscal_period_list")
        history_url = reverse(
            "finance:fiscal_year_history",
            args=[self.fiscal_year.pk],
        )
        self.client.force_login(self.accountant)

        list_response = self.client.get(list_url)
        self.assertContains(list_response, history_url)

        first_page = self.client.get(history_url)
        self.assertEqual(first_page.status_code, 200)
        self.assertEqual(len(first_page.context["page_obj"]), 10)
        self.assertTrue(first_page.context["page_obj"].has_next())

        second_page = self.client.get(history_url, {"page": 2})
        self.assertEqual(len(second_page.context["page_obj"]), 1)
        self.assertTrue(second_page.context["page_obj"].has_previous())

    def test_cashier_cannot_view_fiscal_history(self):
        self.client.force_login(self.cashier)
        response = self.client.get(
            reverse(
                "finance:fiscal_year_history",
                args=[self.fiscal_year.pk],
            )
        )
        self.assertEqual(response.status_code, 403)


class PartyStatementTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="party-statement-accountant",
            password="test-password",
            role="accountant",
        )
        self.receivable = Account.objects.create(
            code="1300",
            name="Accounts receivable",
            account_type=Account.AccountType.ASSET,
        )
        self.payable = Account.objects.create(
            code="2100",
            name="Accounts payable",
            account_type=Account.AccountType.LIABILITY,
        )
        self.customer = Customer.objects.create(
            code="ST-C001",
            name="Statement customer",
            opening_balance=Decimal("25.000"),
        )
        self.supplier = Supplier.objects.create(
            code="ST-S001",
            name="Statement supplier",
            opening_balance=Decimal("30.000"),
        )

    def add_line(
        self,
        number,
        journal_date,
        account,
        *,
        debit="0.000",
        credit="0.000",
        customer=None,
        supplier=None,
        status=JournalEntry.Status.POSTED,
    ):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=journal_date,
            description=f"Activity for {number}",
            status=status,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        return JournalEntryLine.objects.create(
            journal_entry=journal,
            account=account,
            customer=customer,
            supplier=supplier,
            debit=Decimal(debit),
            credit=Decimal(credit),
        )

    def test_customer_statement_uses_opening_and_debit_normal_balance(self):
        self.add_line(
            "CUST-OLD",
            date(2026, 1, 15),
            self.receivable,
            debit="100.000",
            customer=self.customer,
        )
        self.add_line(
            "CUST-PAY",
            date(2026, 2, 10),
            self.receivable,
            credit="40.000",
            customer=self.customer,
        )
        self.add_line(
            "CUST-DRAFT",
            date(2026, 2, 11),
            self.receivable,
            debit="999.000",
            customer=self.customer,
            status=JournalEntry.Status.DRAFT,
        )

        statement = generate_customer_statement(
            self.customer,
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
        )

        self.assertEqual(statement["opening_balance"], Decimal("125.000"))
        self.assertEqual(len(statement["entries"]), 1)
        self.assertEqual(statement["entries"][0]["balance"], Decimal("85.000"))
        self.assertEqual(statement["closing_balance"], Decimal("85.000"))

    def test_supplier_statement_uses_opening_and_credit_normal_balance(self):
        self.add_line(
            "SUP-OLD",
            date(2026, 1, 20),
            self.payable,
            credit="80.000",
            supplier=self.supplier,
        )
        self.add_line(
            "SUP-PAY",
            date(2026, 2, 12),
            self.payable,
            debit="25.000",
            supplier=self.supplier,
        )

        statement = generate_supplier_statement(
            self.supplier,
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
        )

        self.assertEqual(statement["opening_balance"], Decimal("110.000"))
        self.assertEqual(len(statement["entries"]), 1)
        self.assertEqual(statement["entries"][0]["balance"], Decimal("85.000"))
        self.assertEqual(statement["closing_balance"], Decimal("85.000"))

    def test_statement_pages_render_and_validate_date_ranges(self):
        self.client.force_login(self.accountant)
        customer_url = reverse("finance:customer_statement")
        supplier_url = reverse("finance:supplier_statement")

        customer_response = self.client.get(
            customer_url,
            {"customer": self.customer.pk},
        )
        self.assertEqual(customer_response.status_code, 200)
        self.assertContains(customer_response, "Statement customer")

        supplier_response = self.client.get(
            supplier_url,
            {"supplier": self.supplier.pk},
        )
        self.assertEqual(supplier_response.status_code, 200)
        self.assertContains(supplier_response, "Statement supplier")

        invalid_response = self.client.get(
            customer_url,
            {
                "customer": self.customer.pk,
                "start_date": "2026-03-01",
                "end_date": "2026-02-01",
            },
        )
        self.assertContains(
            invalid_response,
            "The start date cannot be later than the end date.",
        )


class AgingReportTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="aging-accountant",
            password="test-password",
            role="accountant",
        )
        self.cashier = get_user_model().objects.create_user(
            username="aging-cashier",
            password="test-password",
            role="cashier",
        )
        self.receivable = Account.objects.create(
            code="1300",
            name="Accounts receivable",
            account_type=Account.AccountType.ASSET,
        )
        self.payable = Account.objects.create(
            code="2100",
            name="Accounts payable",
            account_type=Account.AccountType.LIABILITY,
        )
        self.cash = Account.objects.create(
            code="1100",
            name="Cash",
            account_type=Account.AccountType.ASSET,
            is_cash_equivalent=True,
        )
        self.customer = Customer.objects.create(
            code="AGE-C001",
            name="Aging customer",
        )
        self.supplier = Supplier.objects.create(
            code="AGE-S001",
            name="Aging supplier",
        )

    def add_party_line(
        self,
        number,
        journal_date,
        *,
        account,
        debit="0.000",
        credit="0.000",
        customer=None,
        supplier=None,
        status=JournalEntry.Status.POSTED,
        source_type=JournalEntry.SourceType.MANUAL,
        source_id=None,
    ):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=journal_date,
            description=number,
            status=status,
            source_type=source_type,
            source_id=source_id,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        return JournalEntryLine.objects.create(
            journal_entry=journal,
            account=account,
            customer=customer,
            supplier=supplier,
            debit=Decimal(debit),
            credit=Decimal(credit),
        )

    def test_receivables_aging_bucket_boundaries_and_fifo_receipt(self):
        for number, journal_date, amount in (
            ("AR-90", date(2026, 5, 31), "50.000"),
            ("AR-61", date(2026, 7, 1), "40.000"),
            ("AR-31", date(2026, 7, 31), "30.000"),
            ("AR-30", date(2026, 8, 1), "20.000"),
        ):
            self.add_party_line(
                number,
                journal_date,
                account=self.receivable,
                debit=amount,
                customer=self.customer,
            )
        self.add_party_line(
            "AR-RECEIPT",
            date(2026, 8, 15),
            account=self.receivable,
            credit="15.000",
            customer=self.customer,
            source_type=JournalEntry.SourceType.RECEIPT,
            source_id=501,
        )
        self.add_party_line(
            "AR-CURRENT",
            date(2026, 8, 31),
            account=self.receivable,
            debit="10.000",
            customer=self.customer,
        )
        self.add_party_line(
            "AR-DRAFT",
            date(2026, 8, 1),
            account=self.receivable,
            debit="999.000",
            customer=self.customer,
            status=JournalEntry.Status.DRAFT,
        )

        report = generate_receivables_aging(
            as_of_date=date(2026, 8, 31),
            customer=self.customer,
        )
        row = report["rows"][0]

        self.assertEqual(row["current"], Decimal("10.000"))
        self.assertEqual(row["days_1_30"], Decimal("20.000"))
        self.assertEqual(row["days_31_60"], Decimal("30.000"))
        self.assertEqual(row["days_61_90"], Decimal("40.000"))
        self.assertEqual(row["days_90_plus"], Decimal("35.000"))
        self.assertEqual(row["balance"], Decimal("135.000"))

    def test_payables_aging_uses_purchase_due_date_and_partial_payment(self):
        warehouse = Warehouse.objects.create(code="AGE-WH", name="Aging warehouse")
        invoice = PurchaseInvoice.objects.create(
            invoice_number="AGE-PI-1",
            supplier=self.supplier,
            warehouse=warehouse,
            invoice_date=date(2026, 1, 1),
            due_date=date(2026, 9, 15),
            payment_type=PurchaseInvoice.PaymentType.CREDIT,
            status=PurchaseInvoice.Status.CONFIRMED,
            total=Decimal("100.000"),
            created_by=self.accountant,
        )
        self.add_party_line(
            "AGE-PI-JOURNAL",
            invoice.invoice_date,
            account=self.payable,
            credit="100.000",
            supplier=self.supplier,
            source_type=JournalEntry.SourceType.PURCHASE,
            source_id=invoice.pk,
        )
        self.add_party_line(
            "AGE-PAYMENT",
            date(2026, 2, 1),
            account=self.payable,
            debit="40.000",
            supplier=self.supplier,
            source_type=JournalEntry.SourceType.PAYMENT,
            source_id=502,
        )

        report = generate_payables_aging(
            as_of_date=date(2026, 8, 31),
            supplier=self.supplier,
        )
        row = report["rows"][0]

        self.assertEqual(row["current"], Decimal("60.000"))
        self.assertEqual(row["total_outstanding"], Decimal("60.000"))
        self.assertEqual(row["documents"][0]["due_date"], date(2026, 9, 15))

    def test_reversed_documents_and_their_reversal_do_not_affect_aging(self):
        self.add_party_line(
            "AR-REVERSED",
            date(2026, 8, 1),
            account=self.receivable,
            debit="70.000",
            customer=self.customer,
            status=JournalEntry.Status.REVERSED,
        )
        self.add_party_line(
            "AR-REVERSAL",
            date(2026, 8, 2),
            account=self.receivable,
            credit="70.000",
            customer=self.customer,
            source_type=JournalEntry.SourceType.REVERSAL,
            source_id=999,
        )

        report = generate_receivables_aging(
            as_of_date=date(2026, 8, 31),
            customer=self.customer,
        )

        self.assertEqual(report["rows"], [])
        self.assertEqual(report["totals"]["balance"], Decimal("0.000"))

    def test_receipt_posting_records_fifo_open_item_allocation(self):
        create_required_fiscal_years(date.today().year)
        source_line = self.add_party_line(
            "AR-OPEN-ITEM",
            date.today(),
            account=self.receivable,
            debit="75.000",
            customer=self.customer,
        )
        open_item = OpenItem.objects.create(
            item_type=OpenItem.ItemType.RECEIVABLE,
            customer=self.customer,
            journal_line=source_line,
            document_number="AR-OPEN-ITEM",
            document_date=date.today(),
            due_date=date.today(),
            original_amount=Decimal("75.000"),
        )
        receipt = ReceiptVoucher.objects.create(
            voucher_number="AGE-RV-1",
            date=date.today(),
            customer=self.customer,
            received_from=self.customer.name,
            account=self.cash,
            amount=Decimal("30.000"),
            payment_method=ReceiptVoucher.PaymentMethod.CASH,
            created_by=self.accountant,
        )

        post_receipt_voucher(receipt.pk, self.accountant)

        allocation = OpenItemAllocation.objects.get(open_item=open_item)
        self.assertEqual(allocation.amount, Decimal("30.000"))
        self.assertEqual(allocation.allocation_date, date.today())

        receipt_journal = JournalEntry.objects.get(
            source_type=JournalEntry.SourceType.RECEIPT,
            source_id=receipt.pk,
        )
        reversal = reverse_journal_entry(
            receipt_journal.pk,
            self.accountant,
            "Receipt entered incorrectly",
        )
        self.assertFalse(
            OpenItemAllocation.objects.filter(open_item=open_item).exists()
        )
        self.assertEqual(
            reversal.lines.get(account=self.receivable).customer,
            self.customer,
        )

    def test_aging_pages_filter_validate_and_enforce_permissions(self):
        self.client.force_login(self.accountant)
        receivables_url = reverse("finance:receivables_aging")
        payables_url = reverse("finance:payables_aging")

        response = self.client.get(
            receivables_url,
            {"as_of_date": "2026-08-31", "customer": self.customer.pk},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Accounts receivable aging")

        invalid = self.client.get(receivables_url, {"as_of_date": "invalid"})
        self.assertContains(invalid, "Enter a valid date.")

        self.assertEqual(self.client.get(payables_url).status_code, 200)

        self.client.force_login(self.cashier)
        self.assertEqual(self.client.get(receivables_url).status_code, 403)
        self.assertEqual(self.client.get(payables_url).status_code, 403)


class CashFlowStatementTests(TestCase):
    def setUp(self):
        self.accountant = get_user_model().objects.create_user(
            username="cashflow-accountant",
            password="test-password",
            role="accountant",
        )
        self.cashier = get_user_model().objects.create_user(
            username="cashflow-cashier",
            password="test-password",
            role="cashier",
        )
        create_required_fiscal_years(2026, date.today().year)
        self.cash = Account.objects.create(
            code="CF1100",
            name="Cash-flow cash",
            account_type=Account.AccountType.ASSET,
            is_cash_equivalent=True,
        )
        self.counterpart = Account.objects.create(
            code="CF3000",
            name="Cash-flow counterpart",
            account_type=Account.AccountType.EQUITY,
        )

    def create_cash_journal(
        self,
        number,
        journal_date,
        amount,
        activity,
        *,
        inflow=True,
        status=JournalEntry.Status.POSTED,
    ):
        journal = JournalEntry.objects.create(
            entry_number=number,
            date=journal_date,
            description=number,
            cash_flow_activity=activity,
            status=status,
            created_by=self.accountant,
            approved_by=(
                self.accountant if status == JournalEntry.Status.POSTED else None
            ),
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.cash,
            debit=amount if inflow else Decimal("0.000"),
            credit=Decimal("0.000") if inflow else amount,
        )
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=self.counterpart,
            debit=Decimal("0.000") if inflow else amount,
            credit=amount if inflow else Decimal("0.000"),
        )
        return journal

    def test_cash_flow_groups_activities_and_reconciles_cash(self):
        self.create_cash_journal(
            "CF-OPEN",
            date(2026, 1, 1),
            Decimal("100.000"),
            JournalEntry.CashFlowActivity.FINANCING,
        )
        self.create_cash_journal(
            "CF-OPERATING-IN",
            date(2026, 2, 5),
            Decimal("50.000"),
            JournalEntry.CashFlowActivity.OPERATING,
        )
        self.create_cash_journal(
            "CF-OPERATING-OUT",
            date(2026, 2, 6),
            Decimal("20.000"),
            JournalEntry.CashFlowActivity.OPERATING,
            inflow=False,
        )
        self.create_cash_journal(
            "CF-INVESTING",
            date(2026, 2, 7),
            Decimal("30.000"),
            JournalEntry.CashFlowActivity.INVESTING,
            inflow=False,
        )
        self.create_cash_journal(
            "CF-FINANCING",
            date(2026, 2, 8),
            Decimal("40.000"),
            JournalEntry.CashFlowActivity.FINANCING,
        )
        self.create_cash_journal(
            "CF-DRAFT",
            date(2026, 2, 9),
            Decimal("999.000"),
            JournalEntry.CashFlowActivity.OPERATING,
            status=JournalEntry.Status.DRAFT,
        )

        statement = generate_cash_flow_statement(
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
        )

        self.assertEqual(statement["opening_cash"], Decimal("100.000"))
        self.assertEqual(statement["operating_total"], Decimal("30.000"))
        self.assertEqual(statement["investing_total"], Decimal("-30.000"))
        self.assertEqual(statement["financing_total"], Decimal("40.000"))
        self.assertEqual(statement["total_inflows"], Decimal("90.000"))
        self.assertEqual(statement["total_outflows"], Decimal("50.000"))
        self.assertEqual(statement["net_change"], Decimal("40.000"))
        self.assertEqual(statement["closing_cash"], Decimal("140.000"))
        self.assertTrue(statement["is_reconciled"])

    def test_cash_journal_requires_activity_before_posting(self):
        journal = self.create_cash_journal(
            "CF-VALIDATION",
            date.today(),
            Decimal("25.000"),
            JournalEntry.CashFlowActivity.NONE,
            status=JournalEntry.Status.DRAFT,
        )

        with self.assertRaisesMessage(ValidationError, "cash-flow activity"):
            post_journal_entry(journal.pk, self.accountant)

        journal.cash_flow_activity = JournalEntry.CashFlowActivity.INVESTING
        journal.save(update_fields=["cash_flow_activity"])
        post_journal_entry(journal.pk, self.accountant)
        journal.refresh_from_db()
        self.assertEqual(journal.status, JournalEntry.Status.POSTED)

    def test_reversal_cancels_cash_flow_and_keeps_classification(self):
        original = self.create_cash_journal(
            "CF-REVERSE",
            date.today(),
            Decimal("35.000"),
            JournalEntry.CashFlowActivity.OPERATING,
        )

        reversal = reverse_journal_entry(
            original.pk,
            self.accountant,
            "Incorrect cash receipt",
        )
        statement = generate_cash_flow_statement(
            start_date=date.today(),
            end_date=date.today(),
        )

        self.assertEqual(
            reversal.cash_flow_activity,
            JournalEntry.CashFlowActivity.OPERATING,
        )
        self.assertEqual(statement["operating_total"], Decimal("0.000"))
        self.assertEqual(statement["net_change"], Decimal("0.000"))

    def test_cash_flow_page_validates_dates_and_permissions(self):
        url = reverse("finance:cash_flow_statement")
        self.client.force_login(self.accountant)

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cash-flow statement")

        invalid = self.client.get(
            url,
            {"start_date": "2026-03-01", "end_date": "2026-02-01"},
        )
        self.assertContains(
            invalid,
            "The start date cannot be later than the end date.",
        )

        self.client.force_login(self.cashier)
        self.assertEqual(self.client.get(url).status_code, 403)

    def test_unclassified_entry_links_to_journal_detail(self):
        journal = self.create_cash_journal(
            "CF-UNCLASSIFIED",
            date.today(),
            Decimal("15.000"),
            JournalEntry.CashFlowActivity.NONE,
        )
        self.client.force_login(self.accountant)

        response = self.client.get(reverse("finance:cash_flow_statement"))

        detail_url = reverse("finance:journal_detail", args=[journal.pk])
        self.assertContains(response, f'href="{detail_url}"')
        self.assertContains(response, "CF-UNCLASSIFIED")
