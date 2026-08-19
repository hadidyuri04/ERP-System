from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.formats import date_format
from django.utils.translation import get_language, gettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Error as PlaywrightError, sync_playwright


EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
MONEY_FORMAT = "#,##0.000;[Red]-#,##0.000"


def _display(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return date_format(value, "SHORT_DATE_FORMAT")
    if isinstance(value, (str, bool, int, float, Decimal)):
        return value
    return str(value)


def _safe_sheet_title(title):
    for character in "[]:*?/\\":
        title = title.replace(character, "-")
    return title[:31] or "Report"


def excel_response(filename, document):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(str(document["title"]))
    worksheet.sheet_view.rightToLeft = get_language() == "ar"
    worksheet.freeze_panes = "A4"

    max_columns = max(
        (len(section["headers"]) for section in document["sections"]),
        default=1,
    )
    worksheet.append([str(document["title"])])
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max_columns,
    )
    worksheet["A1"].font = Font(size=16, bold=True, color="2C2341")
    worksheet["A1"].alignment = Alignment(horizontal="center")

    if document.get("metadata"):
        metadata_text = " | ".join(
            f"{label}: {_display(value)}"
            for label, value in document["metadata"]
            if value not in (None, "")
        )
        worksheet.append([metadata_text])
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=max_columns,
        )
        worksheet["A2"].alignment = Alignment(horizontal="center")
        worksheet["A2"].font = Font(size=10, color="746D83")
    else:
        worksheet.append([])

    worksheet.append([])
    for index, section in enumerate(document["sections"]):
        if index:
            worksheet.append([])

        if section.get("title"):
            worksheet.append([str(section["title"])])
            row_number = worksheet.max_row
            worksheet.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=len(section["headers"]),
            )
            cell = worksheet.cell(row=row_number, column=1)
            cell.font = Font(size=12, bold=True, color="6547E8")

        worksheet.append([str(header) for header in section["headers"]])
        header_row = worksheet.max_row
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="6547E8")
            cell.alignment = Alignment(horizontal="center")

        for values in section.get("rows", []):
            worksheet.append([_display(value) for value in values])
            for cell in worksheet[worksheet.max_row]:
                if isinstance(cell.value, (Decimal, float, int)):
                    cell.number_format = MONEY_FORMAT

        for values in section.get("footer_rows", []):
            worksheet.append([_display(value) for value in values])
            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True, color="2C2341")
                cell.fill = PatternFill("solid", fgColor="F2EEFF")
                if isinstance(cell.value, (Decimal, float, int)):
                    cell.number_format = MONEY_FORMAT

    for column_index in range(1, max_columns + 1):
        values = [
            str(worksheet.cell(row=row, column=column_index).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max((len(value) for value in values), default=0) + 3, 12),
            42,
        )

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def pdf_response(request, filename, document):
    html = render_to_string(
        "reports/pdf/report.html",
        {"document": document, "language_code": get_language()},
        request=request,
    )

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.set_content(html, wait_until="load")
            pdf = page.pdf(
                format="A4",
                print_background=True,
                prefer_css_page_size=True,
                margin={
                    "top": "12mm",
                    "right": "10mm",
                    "bottom": "14mm",
                    "left": "10mm",
                },
            )
        finally:
            browser.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export_response(request, export_type, filename, document):
    if export_type == "xlsx":
        return excel_response(filename, document)
    if export_type == "pdf":
        try:
            return pdf_response(request, filename, document)
        except PlaywrightError:
            messages.error(
                request,
                _(
                    "PDF export is unavailable. Install the Playwright Chromium "
                    "runtime and try again."
                ),
            )
    return None


def _period_metadata(start_date=None, end_date=None):
    return [
        (_("Start date"), start_date or _("Beginning")),
        (_("End date"), end_date or _("Today")),
    ]


def general_ledger_document(data, start_date=None, end_date=None):
    return {
        "title": _("General ledger"),
        "metadata": [
            (_("Account"), data["account"]),
            *_period_metadata(start_date, end_date),
        ],
        "sections": [{
            "headers": [
                _("Date"), _("Entry"), _("Description"),
                _("Debit"), _("Credit"), _("Running balance"),
            ],
            "rows": [
                [
                    row["date"], row["entry_number"], row["description"],
                    row["debit"], row["credit"], row["running_balance"],
                ]
                for row in data["entries"]
            ],
            "footer_rows": [["", "", _("Final balance"), "", "", data["final_balance"]]],
        }],
    }


def trial_balance_document(data, start_date=None, end_date=None):
    return {
        "title": _("Trial balance"),
        "metadata": _period_metadata(start_date, end_date),
        "sections": [{
            "headers": [_("Code"), _("Account"), _("Debit"), _("Credit")],
            "rows": [
                [
                    row["account_code"], row["account_name"],
                    row["total_debit"], row["total_credit"],
                ]
                for row in data["rows"]
            ],
            "footer_rows": [["", _("Total"), data["total_debit"], data["total_credit"]]],
        }],
    }


def income_statement_document(data, start_date=None, end_date=None):
    return {
        "title": _("Income statement"),
        "metadata": _period_metadata(start_date, end_date),
        "sections": [
            {
                "title": _("Revenue"),
                "headers": [_("Code"), _("Account"), _("Amount")],
                "rows": [[row["account_code"], row["account_name"], row["amount"]] for row in data["revenue_rows"]],
                "footer_rows": [["", _("Total revenue"), data["total_revenue"]]],
            },
            {
                "title": _("Expenses"),
                "headers": [_("Code"), _("Account"), _("Amount")],
                "rows": [[row["account_code"], row["account_name"], row["amount"]] for row in data["expense_rows"]],
                "footer_rows": [
                    ["", _("Total expenses"), data["total_expenses"]],
                    ["", _("Net profit"), data["net_profit"]],
                ],
            },
        ],
    }


def balance_sheet_document(data, as_of_date=None):
    account_headers = [_("Code"), _("Account"), _("Amount")]
    return {
        "title": _("Balance sheet"),
        "metadata": [(_("As of date"), as_of_date or _("Today"))],
        "sections": [
            {
                "title": _("Assets"),
                "headers": account_headers,
                "rows": [[row["account_code"], row["account_name"], row["amount"]] for row in data["asset_rows"]],
                "footer_rows": [["", _("Total assets"), data["total_assets"]]],
            },
            {
                "title": _("Liabilities"),
                "headers": account_headers,
                "rows": [[row["account_code"], row["account_name"], row["amount"]] for row in data["liability_rows"]],
                "footer_rows": [["", _("Total liabilities"), data["total_liabilities"]]],
            },
            {
                "title": _("Equity"),
                "headers": account_headers,
                "rows": [[row["account_code"], row["account_name"], row["amount"]] for row in data["equity_rows"]],
                "footer_rows": [
                    ["", _("Current earnings"), data["current_earnings"]],
                    ["", _("Total equity and earnings"), data["total_equity_and_earnings"]],
                    ["", _("Liabilities and equity"), data["total_liabilities_and_equity"]],
                ],
            },
        ],
    }


def cash_flow_document(data):
    headers = [
        _("Date"), _("Entry"), _("Description"), _("Account"),
        _("Inflow"), _("Outflow"), _("Net cash flow"),
    ]

    def section(title, rows, total):
        return {
            "title": title,
            "headers": headers,
            "rows": [
                [
                    row["date"], row["entry_number"], row["description"],
                    row["account"], row["inflow"], row["outflow"], row["net"],
                ]
                for row in rows
            ],
            "footer_rows": [["", "", "", _("Activity total"), "", "", total]],
        }

    sections = [
        section(_("Operating activities"), data["operating_rows"], data["operating_total"]),
        section(_("Investing activities"), data["investing_rows"], data["investing_total"]),
        section(_("Financing activities"), data["financing_rows"], data["financing_total"]),
    ]
    if data["unclassified_rows"]:
        sections.append(section(
            _("Unclassified activity"),
            data["unclassified_rows"],
            data["unclassified_total"],
        ))
    sections.append({
        "title": _("Cash reconciliation"),
        "headers": [_("Measure"), _("Amount")],
        "rows": [
            [_("Opening cash"), data["opening_cash"]],
            [_("Total inflows"), data["total_inflows"]],
            [_("Total outflows"), data["total_outflows"]],
            [_("Net change in cash"), data["net_change"]],
        ],
        "footer_rows": [[_("Closing cash"), data["closing_cash"]]],
    })
    return {
        "title": _("Cash-flow statement"),
        "metadata": _period_metadata(data.get("start_date"), data.get("end_date")),
        "sections": sections,
    }


def party_statement_document(data, kind, start_date=None, end_date=None):
    title = _("Customer statement") if kind == "customer" else _("Supplier statement")
    return {
        "title": title,
        "metadata": [
            (_("Account"), data["party"]),
            (_("Opening balance"), data["opening_balance"]),
            *_period_metadata(start_date, end_date),
        ],
        "sections": [{
            "headers": [
                _("Date"), _("Entry"), _("Description"),
                _("Debit"), _("Credit"), _("Balance"),
            ],
            "rows": [
                [
                    row["date"], row["entry_number"], row["description"],
                    row["debit"], row["credit"], row["balance"],
                ]
                for row in data["entries"]
            ],
            "footer_rows": [["", "", _("Closing balance"), "", "", data["closing_balance"]]],
        }],
    }


def aging_document(data, kind):
    title = _("Accounts receivable aging") if kind == "receivable" else _("Accounts payable aging")
    party_label = _("Customer") if kind == "receivable" else _("Supplier")
    details = []
    for row in data["rows"]:
        for item in row["documents"]:
            details.append([
                row["party"], item["document_number"], item["document_date"],
                item["due_date"], item["days_overdue"],
                item["original_amount"], item["remaining"],
            ])

    sections = [{
        "title": _("Aging by account"),
        "headers": [
            party_label, _("Current"), _("1-30 days"), _("31-60 days"),
            _("61-90 days"), _("90+ days"), _("Credit"), _("Balance"),
        ],
        "rows": [
            [
                row["party"], row["current"], row["days_1_30"],
                row["days_31_60"], row["days_61_90"], row["days_90_plus"],
                row["unapplied_credit"], row["balance"],
            ]
            for row in data["rows"]
        ],
        "footer_rows": [[
            _("Report total"), data["totals"]["current"],
            data["totals"]["days_1_30"], data["totals"]["days_31_60"],
            data["totals"]["days_61_90"], data["totals"]["days_90_plus"],
            data["totals"]["unapplied_credit"], data["totals"]["balance"],
        ]],
    }]
    if details:
        sections.append({
            "title": _("Open documents"),
            "headers": [
                party_label, _("Document"), _("Document date"), _("Due date"),
                _("Days overdue"), _("Original"), _("Remaining"),
            ],
            "rows": details,
        })
    return {
        "title": title,
        "metadata": [(_("As of date"), data["as_of_date"])],
        "sections": sections,
    }
